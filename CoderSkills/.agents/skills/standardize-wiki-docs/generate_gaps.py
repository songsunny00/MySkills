#!/usr/bin/env python3
"""将缺失信息 JSON 生成待补充清单 Excel。"""

import sys
import json
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


PRIORITY_COLORS = {
    "必填": "FF4444",
    "建议补充": "FF9800",
    "可选": "4CAF50",
}

PRIORITY_ROW_FILLS = {
    "必填": PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid"),
    "建议补充": PatternFill(start_color="FFFBF0", end_color="FFFBF0", fill_type="solid"),
}

SECTION_COLORS = {
    "文档元数据": "E3F2FD",
    "需求概述": "E8F5E9",
    "功能详细说明": "FFF3E0",
    "功能规则": "FCE4EC",
    "界面原型": "F3E5F5",
    "异常处理": "FFEBEE",
    "非功能性需求": "E0F7FA",
    "依赖关系": "FFF8E1",
    "验收标准": "F1F8E9",
    "FAQ": "EDE7F6",
}


def generate_excel(gaps: list[dict], output_path: str):
    """生成待补充清单 Excel。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "待补充清单"

    headers = ["序号", "所属模块", "模板章节", "缺失内容", "优先级", "已有上下文", "补充建议", "产品补充", "状态"]
    col_widths = [6, 15, 16, 35, 10, 40, 35, 35, 12]

    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 设置列宽
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 写数据
    for row_idx, gap in enumerate(gaps, 2):
        values = [
            gap.get("id", row_idx - 1),
            gap.get("module", ""),
            gap.get("section", ""),
            gap.get("field", ""),
            gap.get("priority", ""),
            gap.get("context", ""),
            gap.get("suggestion", ""),
            "",  # 产品补充（留空给产品经理填写）
            "待补充",  # 状态默认值
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

        priority = gap.get("priority", "")

        # 优先级文字着色
        if priority in PRIORITY_COLORS:
            ws.cell(row=row_idx, column=5).font = Font(
                name="微软雅黑", size=10, bold=True, color=PRIORITY_COLORS[priority]
            )

        # 高优先级整行背景色
        if priority in PRIORITY_ROW_FILLS:
            for col_idx in range(1, len(headers) + 1):
                if col_idx != 3:  # 不覆盖章节列的特殊背景色
                    ws.cell(row=row_idx, column=col_idx).fill = PRIORITY_ROW_FILLS[priority]

        # 章节背景色
        section = gap.get("section", "")
        if section in SECTION_COLORS:
            fill = PatternFill(
                start_color=SECTION_COLORS[section],
                end_color=SECTION_COLORS[section],
                fill_type="solid",
            )
            ws.cell(row=row_idx, column=3).fill = fill

    # 状态列下拉选项
    from openpyxl.worksheet.datavalidation import DataValidation
    status_dv = DataValidation(
        type="list",
        formula1='"待补充,已补充,不需要,延期处理"',
        allow_blank=True,
    )
    status_dv.error = "请选择有效的状态"
    status_dv.errorTitle = "无效状态"
    status_col = get_column_letter(9)
    status_dv.add(f"{status_col}2:{status_col}{len(gaps) + 1}")
    ws.add_data_validation(status_dv)

    # 优先级列下拉选项
    priority_dv = DataValidation(
        type="list",
        formula1='"必填,建议补充,可选"',
        allow_blank=True,
    )
    priority_col = get_column_letter(5)
    priority_dv.add(f"{priority_col}2:{priority_col}{len(gaps) + 1}")
    ws.add_data_validation(priority_dv)

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(gaps) + 1}"

    # 添加汇总 sheet
    _add_summary_sheet(wb, gaps)

    wb.save(output_path)
    print(f"待补充清单已生成: {output_path}")
    counts = {"必填": 0, "建议补充": 0, "可选": 0}
    for g in gaps:
        p = g.get("priority", "")
        if p in counts:
            counts[p] += 1
    print(f"共 {len(gaps)} 项 (必填:{counts['必填']}, 建议补充:{counts['建议补充']}, 可选:{counts['可选']})")


def _add_summary_sheet(wb: Workbook, gaps: list[dict]):
    """添加汇总统计 sheet。"""
    ws = wb.create_sheet("汇总统计")
    header_font = Font(name="微软雅黑", bold=True, size=11)
    cell_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_table(start_row: int, title: str, col1_name: str, data: dict) -> int:
        ws.cell(row=start_row, column=1, value=title).font = header_font
        row = start_row + 1
        ws.cell(row=row, column=1, value=col1_name).font = header_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value="数量").font = header_font
        ws.cell(row=row, column=2).border = thin_border
        row += 1
        for key, count in data.items():
            ws.cell(row=row, column=1, value=key).font = cell_font
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=count).font = cell_font
            ws.cell(row=row, column=2).border = thin_border
            row += 1
        return row + 1

    # 按章节统计
    section_counts = {}
    for gap in gaps:
        s = gap.get("section", "未分类")
        section_counts[s] = section_counts.get(s, 0) + 1
    next_row = write_table(1, "按模板章节统计", "章节", section_counts)

    # 按优先级统计
    pri_counts = {}
    for gap in gaps:
        p = gap.get("priority", "未分类")
        pri_counts[p] = pri_counts.get(p, 0) + 1
    next_row = write_table(next_row, "按优先级统计", "优先级", pri_counts)

    # 按模块统计
    mod_counts = {}
    for gap in gaps:
        m = gap.get("module", "未分类")
        mod_counts[m] = mod_counts.get(m, 0) + 1
    write_table(next_row, "按模块统计", "模块", mod_counts)

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10


def main():
    if len(sys.argv) < 2:
        print("用法: python generate_gaps.py <JSON文件路径> [输出Excel路径]")
        sys.exit(1)

    json_path = Path(sys.argv[1])
    if not json_path.exists():
        print(f"JSON 文件不存在: {json_path}")
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        gaps = json.load(f)

    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        date_str = datetime.now().strftime("%Y%m%d")
        output_path = str(json_path.parent / f"待补充清单_{date_str}.xlsx")

    generate_excel(gaps, output_path)


if __name__ == "__main__":
    main()
