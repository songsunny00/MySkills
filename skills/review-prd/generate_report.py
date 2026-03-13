#!/usr/bin/env python3
"""将审查结果 JSON 生成 Excel 报告。"""

import sys
import json
from pathlib import Path
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("需要安装 openpyxl: pip install openpyxl")
    sys.exit(1)


SEVERITY_COLORS = {
    "高": "FF4444",
    "中": "FF9800",
    "低": "4CAF50",
}

SEVERITY_ROW_FILLS = {
    "高": PatternFill(start_color="FFF0F0", end_color="FFF0F0", fill_type="solid"),
    "中": PatternFill(start_color="FFFBF0", end_color="FFFBF0", fill_type="solid"),
}

TYPE_COLORS = {
    "需求清晰度": "E3F2FD",
    "交互场景完整性": "FFF3E0",
    "逻辑一致性": "FCE4EC",
    "异常处理覆盖": "F3E5F5",
}


def generate_excel(issues: list[dict], output_path: str):
    """生成 Excel 审查报告。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "PRD审查报告"

    headers = ["序号", "所属模块", "问题类型", "严重程度", "问题描述", "文档来源", "修改建议", "产品回复", "状态"]
    col_widths = [6, 15, 16, 10, 50, 25, 40, 30, 12]

    header_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2196F3", end_color="2196F3", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="微软雅黑", size=10)
    cell_align = Alignment(vertical="top", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # 写表头
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 设置列宽
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 写数据
    for row_idx, issue in enumerate(issues, 2):
        values = [
            issue.get("id", row_idx - 1),
            issue.get("module", ""),
            issue.get("type", ""),
            issue.get("severity", ""),
            issue.get("description", ""),
            issue.get("source", ""),
            issue.get("suggestion", ""),
            "",  # 产品回复（留空给产品填写）
            "待确认",  # 状态默认值
        ]

        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

        severity = issue.get("severity", "")

        # 严重程度文字着色
        if severity in SEVERITY_COLORS:
            ws.cell(row=row_idx, column=4).font = Font(
                name="微软雅黑", size=10, bold=True, color=SEVERITY_COLORS[severity]
            )

        # 高/中严重度整行背景色
        if severity in SEVERITY_ROW_FILLS:
            for col_idx in range(1, len(headers) + 1):
                # 不覆盖问题类型列的特殊背景色
                if col_idx != 3:
                    ws.cell(row=row_idx, column=col_idx).fill = SEVERITY_ROW_FILLS[severity]

        # 问题类型背景色
        issue_type = issue.get("type", "")
        if issue_type in TYPE_COLORS:
            fill = PatternFill(
                start_color=TYPE_COLORS[issue_type],
                end_color=TYPE_COLORS[issue_type],
                fill_type="solid",
            )
            ws.cell(row=row_idx, column=3).fill = fill

    # 状态列下拉选项（数据验证）
    from openpyxl.worksheet.datavalidation import DataValidation
    status_dv = DataValidation(
        type="list",
        formula1='"待确认,已确认,已修复,不修改,延期处理"',
        allow_blank=True,
    )
    status_dv.error = "请选择有效的状态"
    status_dv.errorTitle = "无效状态"
    status_col = get_column_letter(9)  # 状态列
    status_dv.add(f"{status_col}2:{status_col}{len(issues) + 1}")
    ws.add_data_validation(status_dv)

    # 冻结首行 + 自动筛选
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(issues) + 1}"

    # 添加汇总 sheet
    _add_summary_sheet(wb, issues)

    wb.save(output_path)
    print(f"报告已生成: {output_path}")
    print(f"共 {len(issues)} 个问题 (高:{sum(1 for i in issues if i.get('severity')=='高')}, "
          f"中:{sum(1 for i in issues if i.get('severity')=='中')}, "
          f"低:{sum(1 for i in issues if i.get('severity')=='低')})")


def _add_summary_sheet(wb: Workbook, issues: list[dict]):
    """添加汇总统计 sheet。"""
    ws = wb.create_sheet("汇总统计")
    header_font = Font(name="微软雅黑", bold=True, size=11)
    cell_font = Font(name="微软雅黑", size=10)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def write_table(start_row: int, title: str, col1_name: str, data: dict) -> int:
        ws.cell(row=start_row, column=1, value=title).font = header_font
        row = start_row + 1
        ws.cell(row=row, column=1, value=col1_name).font = header_font
        ws.cell(row=row, column=1).border = thin_border
        ws.cell(row=row, column=2, value="数量").font = header_font
        ws.cell(row=row, column=2).border = thin_border
        row += 1
        for key, count in data.items():
            ws.cell(row=row, column=1, value=key).font = cell_font
            ws.cell(row=row, column=1).border = thin_border
            ws.cell(row=row, column=2, value=count).font = cell_font
            ws.cell(row=row, column=2).border = thin_border
            row += 1
        return row + 1

    # 按问题类型统计
    type_counts = {}
    for issue in issues:
        t = issue.get("type", "未分类")
        type_counts[t] = type_counts.get(t, 0) + 1
    next_row = write_table(1, "按问题类型统计", "问题类型", type_counts)

    # 按严重程度统计
    sev_counts = {}
    for issue in issues:
        s = issue.get("severity", "未分类")
        sev_counts[s] = sev_counts.get(s, 0) + 1
    next_row = write_table(next_row, "按严重程度统计", "严重程度", sev_counts)

    # 按模块统计
    mod_counts = {}
    for issue in issues:
        m = issue.get("module", "未分类")
        mod_counts[m] = mod_counts.get(m, 0) + 1
    next_row = write_table(next_row, "按模块统计", "模块", mod_counts)

    # 按模块×严重程度交叉统计
    ws.cell(row=next_row, column=1, value="按模块×严重程度统计").font = header_font
    row = next_row + 1
    severities = ["高", "中", "低"]
    ws.cell(row=row, column=1, value="模块").font = header_font
    ws.cell(row=row, column=1).border = thin_border
    for i, s in enumerate(severities, 2):
        ws.cell(row=row, column=i, value=s).font = header_font
        ws.cell(row=row, column=i).border = thin_border
    ws.cell(row=row, column=len(severities) + 2, value="合计").font = header_font
    ws.cell(row=row, column=len(severities) + 2).border = thin_border
    row += 1

    cross: dict[str, dict[str, int]] = {}
    for issue in issues:
        m = issue.get("module", "未分类")
        s = issue.get("severity", "未分类")
        cross.setdefault(m, {}).setdefault(s, 0)
        cross[m][s] = cross[m].get(s, 0) + 1

    for mod, sev_map in cross.items():
        ws.cell(row=row, column=1, value=mod).font = cell_font
        ws.cell(row=row, column=1).border = thin_border
        total = 0
        for i, s in enumerate(severities, 2):
            count = sev_map.get(s, 0)
            ws.cell(row=row, column=i, value=count).font = cell_font
            ws.cell(row=row, column=i).border = thin_border
            total += count
        ws.cell(row=row, column=len(severities) + 2, value=total).font = cell_font
        ws.cell(row=row, column=len(severities) + 2).border = thin_border
        row += 1

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10


def main():
    if len(sys.argv) < 2:
        print("用法: python generate_report.py <JSON文件路径> [输出Excel路径]")
        sys.exit(1)

    json_path = Path(sys.argv[1])
    if not json_path.exists():
        print(f"JSON 文件不存在: {json_path}")
        sys.exit(1)

    with open(json_path, "r", encoding="utf-8") as f:
        issues = json.load(f)

    if len(sys.argv) >= 3:
        output_path = sys.argv[2]
    else:
        date_str = datetime.now().strftime("%Y%m%d")
        output_path = str(json_path.parent / f"PRD审查报告_{date_str}.xlsx")

    generate_excel(issues, output_path)


if __name__ == "__main__":
    main()